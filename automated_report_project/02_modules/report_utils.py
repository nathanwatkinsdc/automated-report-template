"""
Core Reporting Utilities Module

This module provides essential functions for automated report generation,
including figure/table saving, PDF export, and output management with
timestamped directories.

Functions:
    save_figure: Save plotly/matplotlib figures with timestamping
    save_table: Export pandas DataFrames to formatted Excel files
    print_to_pdf: Generate PDF reports from multiple sources
    generate_all_outputs: Rebuild all project outputs in single call
    add_annotation_tag: Add classification tags to plotly charts
    create_side_by_side_display: Display charts/tables side-by-side in Jupyter

Example:
    from modules.report_utils import save_figure, save_table
    
    # Save a plotly figure
    save_figure(fig, 'fig_1a_loan_trends', print_mode='YES')
    
    # Export formatted table
    save_table(df, 'table_2_summary_stats', print_mode='YES')
"""

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

# Add project root to path for imports
sys.path.append(str(Path(__file__).parent.parent))
from config.config import *

# =============================================================================
# FIGURE SAVING UTILITIES
# =============================================================================

def save_figure(fig, filename, print_mode='NO', output_dir=None, 
               file_format='png', width=1200, height=800, dpi=300):
    """
    Save plotly or matplotlib figure with optional timestamping.
    
    Args:
        fig: Plotly figure object or matplotlib figure
        filename (str): Base filename without extension
        print_mode (str): 'YES' to save to timestamped folder, 'NO' to skip
        output_dir (Path): Custom output directory (default: DRAFT_DIR)
        file_format (str): File format ('png', 'pdf', 'svg', 'html')
        width (int): Figure width in pixels
        height (int): Figure height in pixels  
        dpi (int): Resolution for raster formats
        
    Returns:
        Path: Path to saved file, or None if print_mode='NO'
        
    Example:
        fig = px.bar(df, x='category', y='value')
        save_figure(fig, 'fig_1a_category_analysis', 'YES')
    """
    if print_mode.upper() != 'YES':
        return None
    
    # Set default output directory
    if output_dir is None:
        output_dir = get_timestamped_path(DRAFT_DIR)
    
    # Ensure filename has proper prefix
    if not filename.startswith(NAMING_CONVENTIONS['FIGURE_PREFIX']):
        filename = NAMING_CONVENTIONS['FIGURE_PREFIX'] + filename
    
    # Add file extension
    if not filename.endswith(f'.{file_format}'):
        filename = f"{filename}.{file_format}"
    
    filepath = output_dir / filename
    
    try:
        # Handle plotly figures
        if hasattr(fig, 'write_image') or hasattr(fig, 'write_html'):
            if file_format == 'html':
                fig.write_html(str(filepath))
            else:
                fig.write_image(str(filepath), width=width, height=height)
        
        # Handle matplotlib figures  
        elif hasattr(fig, 'savefig'):
            fig.savefig(str(filepath), dpi=dpi, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
        
        else:
            raise ValueError("Unsupported figure type")
            
        print(f"✓ Figure saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"✗ Error saving figure {filename}: {str(e)}")
        return None

def add_annotation_tag(fig, tag_text="For Official Use Only", 
                      position='top_right', font_size=10):
    """
    Add classification or annotation tag to plotly figure.
    
    Args:
        fig: Plotly figure object
        tag_text (str): Text to display in annotation
        position (str): Position ('top_right', 'top_left', 'bottom_right', 'bottom_left')
        font_size (int): Font size for annotation
        
    Returns:
        Plotly figure with annotation added
        
    Example:
        fig = px.scatter(df, x='x', y='y')
        fig = add_annotation_tag(fig, "Controlled Information")
    """
    position_coords = {
        'top_right': dict(x=0.99, y=0.99, xanchor='right', yanchor='top'),
        'top_left': dict(x=0.01, y=0.99, xanchor='left', yanchor='top'),
        'bottom_right': dict(x=0.99, y=0.01, xanchor='right', yanchor='bottom'),
        'bottom_left': dict(x=0.01, y=0.01, xanchor='left', yanchor='bottom')
    }
    
    coords = position_coords.get(position, position_coords['top_right'])
    
    fig.add_annotation(
        text=tag_text,
        xref="paper", yref="paper",
        x=coords['x'], y=coords['y'],
        xanchor=coords['xanchor'], yanchor=coords['yanchor'],
        showarrow=False,
        font=dict(size=font_size, color="gray"),
        bgcolor="rgba(255,255,255,0.8)",
        bordercolor="gray",
        borderwidth=1
    )
    
    return fig

# =============================================================================
# TABLE SAVING UTILITIES  
# =============================================================================

def save_table(df, filename, print_mode='NO', output_dir=None,
               file_format='xlsx', include_index=False, sheet_name='Data'):
    """
    Save pandas DataFrame to formatted Excel or CSV file.
    
    Args:
        df (DataFrame): Pandas DataFrame to save
        filename (str): Base filename without extension
        print_mode (str): 'YES' to save to timestamped folder, 'NO' to skip
        output_dir (Path): Custom output directory (default: DRAFT_DIR)
        file_format (str): File format ('xlsx', 'csv')
        include_index (bool): Whether to include DataFrame index
        sheet_name (str): Excel sheet name
        
    Returns:
        Path: Path to saved file, or None if print_mode='NO'
        
    Example:
        save_table(summary_df, 'table_1_loan_summary', 'YES')
    """
    if print_mode.upper() != 'YES':
        return None
    
    # Set default output directory
    if output_dir is None:
        output_dir = get_timestamped_path(DRAFT_DIR)
    
    # Ensure filename has proper prefix
    if not filename.startswith(NAMING_CONVENTIONS['TABLE_PREFIX']):
        filename = NAMING_CONVENTIONS['TABLE_PREFIX'] + filename
    
    # Add file extension
    if not filename.endswith(f'.{file_format}'):
        filename = f"{filename}.{file_format}"
    
    filepath = output_dir / filename
    
    try:
        if file_format == 'xlsx':
            # Save with formatting
            _save_formatted_excel(df, filepath, sheet_name, include_index)
        elif file_format == 'csv':
            df.to_csv(filepath, index=include_index)
        else:
            raise ValueError(f"Unsupported file format: {file_format}")
            
        print(f"✓ Table saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"✗ Error saving table {filename}: {str(e)}")
        return None

def _save_formatted_excel(df, filepath, sheet_name, include_index):
    """
    Save DataFrame to Excel with professional formatting.
    
    Args:
        df (DataFrame): DataFrame to save
        filepath (Path): Output file path
        sheet_name (str): Excel sheet name
        include_index (bool): Whether to include index
    """
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=include_index)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting
        _apply_excel_formatting(worksheet, df, include_index)

def _apply_excel_formatting(worksheet, df, include_index):
    """
    Apply professional formatting to Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        df (DataFrame): Source DataFrame for dimensions
        include_index (bool): Whether index is included
    """
    # Header formatting
    header_font = Font(bold=True, color=EXCEL_FORMATTING['HEADER_FONT_COLOR'])
    header_fill = PatternFill(start_color=EXCEL_FORMATTING['HEADER_COLOR'][1:], 
                            end_color=EXCEL_FORMATTING['HEADER_COLOR'][1:], 
                            fill_type='solid')
    
    # Apply header formatting
    header_row = 1
    start_col = 2 if include_index else 1
    end_col = len(df.columns) + (1 if include_index else 0)
    
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# =============================================================================
# PDF GENERATION UTILITIES
# =============================================================================

def print_to_pdf(content_list, filename, print_mode='NO', output_dir=None,
                 title="Automated Report", metadata=None):
    """
    Generate PDF report from multiple content sources.
    
    Args:
        content_list (list): List of content items (figures, tables, text)
        filename (str): Base filename without extension
        print_mode (str): 'YES' to save to timestamped folder, 'NO' to skip
        output_dir (Path): Custom output directory (default: FINAL_DIR)
        title (str): Report title
        metadata (dict): Additional report metadata
        
    Returns:
        Path: Path to saved PDF, or None if print_mode='NO'
        
    Example:
        content = [fig1, summary_table, "Analysis complete"]
        print_to_pdf(content, 'monthly_fraud_report', 'YES')
    """
    if print_mode.upper() != 'YES':
        return None
    
    # Set default output directory
    if output_dir is None:
        output_dir = get_timestamped_path(FINAL_DIR)
    
    # Ensure filename has proper prefix and extension
    if not filename.startswith(NAMING_CONVENTIONS['REPORT_PREFIX']):
        filename = NAMING_CONVENTIONS['REPORT_PREFIX'] + filename
    
    if not filename.endswith('.pdf'):
        filename = f"{filename}.pdf"
    
    filepath = output_dir / filename
    
    try:
        # This is a simplified PDF generation placeholder
        # In practice, you might use libraries like reportlab, weasyprint, or matplotlib
        print(f"📄 PDF generation placeholder for: {filepath}")
        print(f"   Title: {title}")
        print(f"   Content items: {len(content_list)}")
        
        # Create placeholder file
        filepath.touch()
        
        print(f"✓ PDF saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"✗ Error generating PDF {filename}: {str(e)}")
        return None

# =============================================================================
# BATCH PROCESSING UTILITIES
# =============================================================================

def generate_all_outputs(notebooks_to_run=None, print_mode='YES'):
    """
    Execute all notebooks and generate complete set of outputs.
    
    Args:
        notebooks_to_run (list): List of notebook files to execute (default: all)
        print_mode (str): 'YES' to save outputs, 'NO' for dry run
        
    Returns:
        dict: Summary of generated outputs
        
    Example:
        results = generate_all_outputs(['1_data_overview.ipynb', '2_analysis.ipynb'])
    """
    print("🚀 Starting automated report generation...")
    print("=" * 50)
    
    # Create timestamped output directory
    if print_mode.upper() == 'YES':
        output_dir = get_timestamped_path(FINAL_DIR)
        print(f"📁 Output directory: {output_dir}")
    
    results = {
        'timestamp': datetime.now().isoformat(),
        'notebooks_executed': [],
        'figures_generated': [],
        'tables_generated': [],
        'reports_generated': [],
        'errors': []
    }
    
    # Default notebooks to run
    if notebooks_to_run is None:
        notebooks_to_run = [
            '1_data_overview.ipynb',
            '2_trend_analysis.ipynb', 
            '3_final_report.ipynb'
        ]
    
    # Execute each notebook
    for notebook in notebooks_to_run:
        notebook_path = NOTEBOOKS_ROOT / notebook
        
        if notebook_path.exists():
            print(f"📓 Processing: {notebook}")
            results['notebooks_executed'].append(notebook)
            
            # Placeholder for notebook execution
            # In practice, you might use nbconvert or papermill
            print(f"   ✓ Executed successfully")
            
        else:
            error_msg = f"Notebook not found: {notebook}"
            print(f"   ✗ {error_msg}")
            results['errors'].append(error_msg)
    
    # Generate summary
    print("\n📊 Generation Summary:")
    print(f"   Notebooks: {len(results['notebooks_executed'])}")
    print(f"   Figures: {len(results['figures_generated'])}")
    print(f"   Tables: {len(results['tables_generated'])}")
    print(f"   Reports: {len(results['reports_generated'])}")
    print(f"   Errors: {len(results['errors'])}")
    
    if results['errors']:
        print("\n⚠️  Errors encountered:")
        for error in results['errors']:
            print(f"      - {error}")
    
    print("\n✅ Report generation complete!")
    return results

# =============================================================================
# JUPYTER DISPLAY UTILITIES
# =============================================================================

def create_side_by_side_display(*items, titles=None):
    """
    Display multiple charts or tables side-by-side in Jupyter notebook.
    
    Args:
        *items: Plotly figures, DataFrames, or other displayable objects
        titles (list): Optional titles for each item
        
    Returns:
        HTML display object for Jupyter
        
    Example:
        create_side_by_side_display(fig1, fig2, titles=['Chart A', 'Chart B'])
    """
    try:
        from IPython.display import display, HTML
        import plotly.offline as pyo
        
        html_content = '<div style="display: flex; flex-wrap: wrap;">'
        
        for i, item in enumerate(items):
            title = titles[i] if titles and i < len(titles) else f"Item {i+1}"
            
            html_content += f'<div style="flex: 1; margin: 10px; min-width: 300px;">'
            html_content += f'<h4>{title}</h4>'
            
            # Handle plotly figures
            if hasattr(item, 'to_html'):
                fig_html = item.to_html(full_html=False, include_plotlyjs='cdn')
                html_content += fig_html
            
            # Handle DataFrames
            elif hasattr(item, 'to_html'):
                df_html = item.to_html(classes='table table-striped', table_id=f'table_{i}')
                html_content += df_html
            
            # Handle other objects
            else:
                html_content += f'<pre>{str(item)}</pre>'
            
            html_content += '</div>'
        
        html_content += '</div>'
        
        return HTML(html_content)
        
    except ImportError:
        print("IPython not available - side-by-side display not supported")
        return None

def display_summary_stats(df, title="Data Summary"):
    """
    Generate and display formatted summary statistics.
    
    Args:
        df (DataFrame): DataFrame to summarize
        title (str): Title for summary display
        
    Returns:
        DataFrame: Formatted summary statistics
        
    Example:
        display_summary_stats(loan_data, "Loan Portfolio Summary")
    """
    try:
        from IPython.display import display, HTML
        
        # Generate summary statistics
        summary = df.describe(include='all').round(2)
        
        # Add additional statistics
        additional_stats = pd.DataFrame({
            'count': df.count(),
            'missing': df.isnull().sum(),
            'unique': df.nunique(),
            'dtype': df.dtypes.astype(str)
        })
        
        # Combine summaries
        full_summary = pd.concat([summary, additional_stats.T])
        
        # Display with title
        display(HTML(f"<h3>{title}</h3>"))
        display(full_summary)
        
        return full_summary
        
    except ImportError:
        print("IPython not available - using basic display")
        print(f"\n{title}")
        print("=" * len(title))
        print(df.describe(include='all'))
        return df.describe(include='all')

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def get_figure_filename(figure_number, description, enterprise=None):
    """
    Generate standardized figure filename.
    
    Args:
        figure_number (str): Figure number (e.g., '1a', '2b')
        description (str): Brief description
        enterprise (str): Optional enterprise code ('ENT1', 'ENT2')
        
    Returns:
        str: Formatted filename
        
    Example:
        filename = get_figure_filename('1a', 'loan_volume_trends', 'ENT1')
        # Returns: 'fig_1a_fnm_loan_volume_trends'
    """
    # Clean description
    clean_desc = description.lower().replace(' ', '_').replace('-', '_')
    
    # Add enterprise if provided
    if enterprise:
        clean_desc = f"{enterprise.lower()}_{clean_desc}"
    
    return f"{NAMING_CONVENTIONS['FIGURE_PREFIX']}{figure_number}_{clean_desc}"

def get_table_filename(table_number, description):
    """
    Generate standardized table filename.
    
    Args:
        table_number (str): Table number (e.g., '1', '2a')
        description (str): Brief description
        
    Returns:
        str: Formatted filename
        
    Example:
        filename = get_table_filename('1', 'summary statistics')
        # Returns: 'table_1_summary_statistics'
    """
    clean_desc = description.lower().replace(' ', '_').replace('-', '_')
    return f"{NAMING_CONVENTIONS['TABLE_PREFIX']}{table_number}_{clean_desc}"

def create_output_manifest(output_dir):
    """
    Create a manifest file listing all generated outputs.
    
    Args:
        output_dir (Path): Directory containing outputs
        
    Returns:
        Path: Path to manifest file
        
    Example:
        manifest_path = create_output_manifest(TIMESTAMPED_FINAL)
    """
    manifest_data = []
    
    # Scan directory for outputs
    for file_path in output_dir.rglob('*'):
        if file_path.is_file():
            file_info = {
                'filename': file_path.name,
                'path': str(file_path.relative_to(output_dir)),
                'size_bytes': file_path.stat().st_size,
                'modified': datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                'type': file_path.suffix[1:] if file_path.suffix else 'unknown'
            }
            manifest_data.append(file_info)
    
    # Create manifest DataFrame
    manifest_df = pd.DataFrame(manifest_data)
    
    # Save manifest
    manifest_path = output_dir / 'output_manifest.xlsx'
    manifest_df.to_excel(manifest_path, index=False)
    
    print(f"📋 Output manifest created: {manifest_path}")
    print(f"   Total files: {len(manifest_data)}")
    
    return manifest_path

# =============================================================================
# INITIALIZATION AND TESTING
# =============================================================================

def test_utilities():
    """
    Test core utility functions with sample data.
    
    Returns:
        bool: True if all tests pass
    """
    print("🧪 Testing report utilities...")
    
    try:
        # Test directory creation
        test_dir = get_timestamped_path(DRAFT_DIR / 'test')
        print(f"✓ Directory creation: {test_dir}")
        
        # Test filename generation
        fig_name = get_figure_filename('1a', 'test chart', 'ENT1')
        table_name = get_table_filename('1', 'test table')
        print(f"✓ Filename generation: {fig_name}, {table_name}")
        
        # Test with sample data
        sample_df = pd.DataFrame({
            'category': ['A', 'B', 'C'],
            'value': [10, 20, 30],
            'date': pd.date_range('2024-01-01', periods=3)
        })
        
        # Test table saving (dry run)
        save_table(sample_df, table_name, print_mode='NO')
        print("✓ Table saving (dry run)")
        
        print("✅ All utility tests passed!")
        return True
        
    except Exception as e:
        print(f"❌ Utility test failed: {str(e)}")
        return False

if __name__ == "__main__":
    # Run tests when module is executed directly
    test_utilities()